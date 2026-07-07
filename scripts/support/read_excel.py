#!/usr/bin/env python3
"""Inspect CSV/XLSX files and optionally export normalized previews."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any

from common import detect_dependency, ensure_dir, exit_if_missing, file_summary, print_report, resolve_mode, utc_timestamp, write_text


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Read CSV/XLSX and produce a structured preview.")
    parser.add_argument("input", type=Path)
    parser.add_argument("--report", action="store_true")
    parser.add_argument("--execute", action="store_true")
    parser.add_argument("--out", type=Path, help="Output directory for summary files.")
    parser.add_argument("--sample-rows", type=int, default=5)
    return parser.parse_args()


def inspect_csv(path: Path, sample_rows: int) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        dialect = csv.Sniffer().sniff(sample) if sample else csv.excel
        reader = csv.reader(handle, dialect)
        rows = list(reader)
    header = rows[0] if rows else []
    return {
        "type": "csv",
        "sheet_count": 1,
        "sheets": [
            {
                "name": path.stem,
                "rows": max(0, len(rows) - 1),
                "columns": len(header),
                "header": header,
                "sample_rows": rows[1 : 1 + sample_rows],
            }
        ],
    }


def inspect_xlsx(path: Path, sample_rows: int) -> dict[str, Any]:
    if not detect_dependency("openpyxl"):
        return {
            "type": "xlsx",
            "dependency_missing": "openpyxl",
            "sheet_count": 0,
            "sheets": [],
        }
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True, max_row=sample_rows + 1))
        header = ["" if cell is None else str(cell) for cell in rows[0]] if rows else []
        samples = [[cell for cell in row] for row in rows[1 : 1 + sample_rows]]
        sheets.append(
            {
                "name": sheet.title,
                "rows": max(sheet.max_row - 1, 0),
                "columns": sheet.max_column,
                "header": header,
                "sample_rows": samples,
            }
        )
    return {
        "type": "xlsx",
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def markdown_summary(path: Path, info: dict[str, Any]) -> str:
    lines = [
        f"Session timestamp: {utc_timestamp()}",
        f"Input: {path}",
        "",
    ]
    for sheet in info.get("sheets", []):
        lines.extend(
            [
                f"## {sheet['name']}",
                f"- Rows: {sheet['rows']}",
                f"- Columns: {sheet['columns']}",
                f"- Header: {sheet['header']}",
                f"- Sample rows: {sheet['sample_rows']}",
                "",
            ]
        )
    return "\n".join(lines)


def main() -> int:
    args = parse_args()
    mode = resolve_mode(args)
    exit_if_missing(args.input)
    ext = args.input.suffix.lower()
    if ext == ".csv":
        info = inspect_csv(args.input, args.sample_rows)
    elif ext == ".xlsx":
        info = inspect_xlsx(args.input, args.sample_rows)
    else:
        raise SystemExit("Unsupported file type. Use .csv or .xlsx")

    report = {
        "mode": mode,
        "session_timestamp": utc_timestamp(),
        "input": file_summary(args.input),
        "inspection": info,
        "will_write": ["summary.md"] if mode == "execute" else [],
    }
    print_report(report)

    if mode == "execute":
        out_dir = args.out or args.input.parent / f"{args.input.stem}-preview"
        ensure_dir(out_dir)
        write_text(out_dir / "summary.md", markdown_summary(args.input, info))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
